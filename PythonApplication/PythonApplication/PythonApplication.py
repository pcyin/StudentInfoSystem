# -*- coding: utf-8 -*-
import openpyxl

def concate_string(cell_col):
    return '\n'.join([ line for line in [' '.join([
                                unicode(cell.value).strip()
                                for cell in row
                                if cell.value != None
                                ]).strip()
                      for row in cell_col]
                      if line != '']).strip()

class Mentor:
    pass

class Student:
    pass

class Education:
    pass

file_name = 'stu.xlsx'
wb = openpyxl.load_workbook(file_name)
ws = wb.get_active_sheet()

stu = Student()

stu.name = ws['C3'].value
stu.no = ws['E3'].value
stu.birthday = ws['G3'].value
stu.hometown = ws['C4'].value
stu.politicstatus = ws['G4'].value + ' ' + str(ws['H4'].value)

stu.hoppy = concate_string(ws['C6':'E6'])
stu.bloodtype = ws['G6'].value
stu.familyinfo = concate_string(ws['B8':'H11'])

stu.eduexp = [ tuple for tuple in 
              [(concate_string(ws['B' + `13+i`:'C' + `13+i`]), 
               concate_string(ws['D' + `13+i`:'E' + `13+i`]), 
               concate_string(ws['F' + `13+i`:'H' + `13+i`]))
              for i in range(1 ,4)]
              if tuple[0] != None]

stu.gaozhong_jingsai = concate_string(ws['B20':'H21'])

stu.gaozhong_xueshenggongzuo = concate_string(ws['B23':'H24'])

stu.gaozhong_qitajiangli = concate_string(ws['B26':'H27'])

stu.benke_chengji = concate_string(ws['B30':'H31'])

stu.benke_jingsai = concate_string(ws['B33':'H34'])

stu.benke_xueshu = concate_string(ws['B36':'H37'])

stu.benke_xueshenggongzuo = concate_string(ws['B39':'H34'])

stu.phd_school = concate_string(ws['C45':'D45'])

stu.phd_direction = concate_string(ws['F45':'H45'])

stu.phd_mentor = Mentor()

stu.phd_mentor.name = concate_string(ws['C46':'D46'])

stu.phd_mentor.phone = concate_string(ws['F46':'H46'])

stu.phd_mentor.email = concate_string(ws['C47':'H47'])

stu.phd_q1 = concate_string(ws['B49':'H51'])
stu.phd_q2 = concate_string(ws['B53':'H54'])

pass
